"""GET /datasets/{id}/data — real-row preview endpoint (stats-refine)."""
from __future__ import annotations

import io

import pytest
from openpyxl import Workbook

XLSX_MIME = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


async def _make_project(client) -> str:
    r = await client.post(
        "/api/projects", json={"title": "Preview", "study_type": "Outcome Study"}
    )
    return r.json()["id"]


async def _upload_csv(client, project_id: str, csv: bytes) -> str:
    files = {"file": ("data.csv", csv, "text/csv")}
    r = await client.post(f"/api/projects/{project_id}/datasets", files=files)
    assert r.status_code == 201, r.text
    return r.json()["id"]


@pytest.mark.asyncio
async def test_preview_returns_actual_rows_not_sample_values(client):
    project_id = await _make_project(client)
    csv = (
        b"age,group\n"
        b"45,A\n50,B\n55,A\n60,B\n40,A\n42,B\n48,A\n52,B\n44,A\n46,B\n"
    )
    ds_id = await _upload_csv(client, project_id, csv)

    r = await client.get(
        f"/api/projects/{project_id}/datasets/{ds_id}/data?offset=0&limit=100"
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["columns"] == ["age", "group"]
    assert body["total"] == 10
    assert len(body["rows"]) == 10
    assert body["rows"][0]["age"] == 45
    assert body["rows"][0]["group"] == "A"
    assert body["rows"][0]["__row_index"] == 0


@pytest.mark.asyncio
async def test_preview_pagination(client):
    project_id = await _make_project(client)
    csv = (b"x\n" + b"\n".join(str(i).encode() for i in range(120)) + b"\n")
    ds_id = await _upload_csv(client, project_id, csv)

    r = await client.get(
        f"/api/projects/{project_id}/datasets/{ds_id}/data?offset=50&limit=20"
    )
    assert r.status_code == 200
    body = r.json()
    assert body["total"] == 120
    assert len(body["rows"]) == 20
    assert body["offset"] == 50
    assert body["rows"][0]["__row_index"] == 50
    assert body["rows"][0]["x"] == 50


@pytest.mark.asyncio
async def test_preview_limit_capped_to_500(client):
    project_id = await _make_project(client)
    csv = b"x\n1\n2\n3\n"
    ds_id = await _upload_csv(client, project_id, csv)
    r = await client.get(
        f"/api/projects/{project_id}/datasets/{ds_id}/data?limit=99999"
    )
    assert r.status_code == 200
    body = r.json()
    # Cap kicks the limit back to the default 50; we have 3 rows total.
    assert body["limit"] == 50


@pytest.mark.asyncio
async def test_preview_404_missing_dataset(client):
    project_id = await _make_project(client)
    r = await client.get(
        f"/api/projects/{project_id}/datasets/missing/data"
    )
    assert r.status_code == 404


@pytest.mark.asyncio
async def test_preview_nan_becomes_null(client):
    project_id = await _make_project(client)
    csv = b"age,group\n45,A\n,B\n55,A\n"
    ds_id = await _upload_csv(client, project_id, csv)
    r = await client.get(
        f"/api/projects/{project_id}/datasets/{ds_id}/data"
    )
    body = r.json()
    assert body["rows"][1]["age"] is None


@pytest.mark.asyncio
async def test_preview_honours_xlsx_sheet_for_multisheet(client):
    """For a multi-sheet upload, each dataset's preview must read its own
    sheet — not the workbook's active sheet."""
    project_id = await _make_project(client)
    wb = Workbook()
    wb.remove(wb.active)
    a = wb.create_sheet("Alpha")
    a.append(["x"])
    a.append([1])
    a.append([2])
    b = wb.create_sheet("Beta")
    b.append(["x"])
    b.append([100])
    b.append([200])
    buf = io.BytesIO()
    wb.save(buf)
    files = {"file": ("mc.xlsx", buf.getvalue(), XLSX_MIME)}
    r = await client.post(f"/api/projects/{project_id}/datasets", files=files)
    assert r.status_code == 201

    r2 = await client.get(f"/api/projects/{project_id}/datasets")
    rows = r2.json()
    by_sheet = {(r["dataset_metadata"] or {}).get("sheet_name"): r["id"] for r in rows}
    alpha_id = by_sheet["Alpha"]
    beta_id = by_sheet["Beta"]

    a_preview = (await client.get(
        f"/api/projects/{project_id}/datasets/{alpha_id}/data"
    )).json()
    b_preview = (await client.get(
        f"/api/projects/{project_id}/datasets/{beta_id}/data"
    )).json()
    assert [r["x"] for r in a_preview["rows"]] == [1, 2]
    assert [r["x"] for r in b_preview["rows"]] == [100, 200]
