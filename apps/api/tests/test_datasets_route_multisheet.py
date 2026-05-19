"""Multi-sheet XLSX upload route — one dataset per sheet."""
from __future__ import annotations

import io

import pytest
from openpyxl import Workbook

XLSX_MIME = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


def _make_multisheet_xlsx() -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    demo = wb.create_sheet("Demographics")
    demo.append(["patient_id", "approach", "age"])
    for i in range(1, 6):
        demo.append([i, "anterior" if i % 2 else "posterior", 40 + i])
    peri = wb.create_sheet("Perioperative")
    peri.append(["patient_id", "operation_time_min"])
    for i in range(1, 6):
        peri.append([i, 60 + i])
    longs = wb.create_sheet("Outcomes_Long")
    longs.append(["patient_id", "timepoint", "hhs"])
    for i in range(1, 6):
        for tp in range(5):
            longs.append([i, tp, 50 + tp * 5])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def _make_project(client, title="Stats Multi") -> str:
    r = await client.post(
        "/api/projects", json={"title": title, "study_type": "Outcome Study"}
    )
    assert r.status_code == 201, r.text
    return r.json()["id"]


@pytest.mark.asyncio
async def test_upload_multisheet_xlsx_creates_one_dataset_per_sheet(client):
    project_id = await _make_project(client)
    data = _make_multisheet_xlsx()
    files = {"file": ("masterchart.xlsx", data, XLSX_MIME)}
    r = await client.post(f"/api/projects/{project_id}/datasets", files=files)
    assert r.status_code == 201, r.text

    r2 = await client.get(f"/api/projects/{project_id}/datasets")
    assert r2.status_code == 200
    rows = r2.json()
    assert len(rows) == 3
    sheet_names = {
        (r["dataset_metadata"] or {}).get("sheet_name") for r in rows
    }
    assert sheet_names == {"Demographics", "Perioperative", "Outcomes_Long"}
    filenames = {r["filename"] for r in rows}
    assert any("Demographics" in fn for fn in filenames)


@pytest.mark.asyncio
async def test_upload_multisheet_xlsx_long_format_hint_on_outcomes(client):
    project_id = await _make_project(client)
    data = _make_multisheet_xlsx()
    files = {"file": ("masterchart.xlsx", data, XLSX_MIME)}
    r = await client.post(f"/api/projects/{project_id}/datasets", files=files)
    assert r.status_code == 201, r.text

    r2 = await client.get(f"/api/projects/{project_id}/datasets")
    rows = r2.json()
    outcomes = next(
        r for r in rows
        if (r["dataset_metadata"] or {}).get("sheet_name") == "Outcomes_Long"
    )
    hint = (outcomes["dataset_metadata"] or {}).get("long_format_hint")
    assert hint is not None
    assert hint["subject_col"] == "patient_id"
    assert hint["time_col"] == "timepoint"


@pytest.mark.asyncio
async def test_upload_single_sheet_xlsx_records_sheet_name(client):
    """A 1-sheet XLSX should still be a single dataset (legacy path) but
    carries sheet_name in dataset_metadata for forward-compat."""
    project_id = await _make_project(client)
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("OnlySheet")
    ws.append(["age", "sex"])
    ws.append([40, "M"])
    ws.append([55, "F"])
    buf = io.BytesIO()
    wb.save(buf)

    files = {"file": ("one.xlsx", buf.getvalue(), XLSX_MIME)}
    r = await client.post(f"/api/projects/{project_id}/datasets", files=files)
    assert r.status_code == 201, r.text
    r2 = await client.get(f"/api/projects/{project_id}/datasets")
    rows = r2.json()
    assert len(rows) == 1
    meta = rows[0]["dataset_metadata"] or {}
    assert meta.get("sheet_name") == "OnlySheet"


@pytest.mark.asyncio
async def test_upload_csv_unchanged(client):
    """CSV path must not have any sheet metadata."""
    project_id = await _make_project(client)
    csv = b"age,group\n45,A\n50,B\n"
    files = {"file": ("data.csv", csv, "text/csv")}
    r = await client.post(f"/api/projects/{project_id}/datasets", files=files)
    assert r.status_code == 201
    body = r.json()
    assert body["dataset_metadata"] is None or "sheet_name" not in (
        body["dataset_metadata"] or {}
    )
