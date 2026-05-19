from __future__ import annotations

import io
import re
from dataclasses import dataclass

import numpy as np
import pandas as pd
from openpyxl import load_workbook

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
CSV_MIME = "text/csv"

_TIME_NAME_RE = re.compile(r"^(date|time|dt|admit|discharge|surgery|fu_|followup)", re.I)

_ORDINAL_TOKENS = {
    "mild",
    "moderate",
    "severe",
    "low",
    "medium",
    "high",
    "small",
    "large",
    "none",
    "minimal",
    "mild-moderate",
    "moderate-severe",
    "i",
    "ii",
    "iii",
    "iv",
    "v",
}


@dataclass(frozen=True)
class InferredColumn:
    name: str
    position: int
    inferred_type: str
    n_missing: int
    sample_values: list[str]


@dataclass(frozen=True)
class IngestResult:
    n_rows: int
    n_columns: int
    columns: list[InferredColumn]
    long_format_hint: dict | None = None


def detect_table_mime(data: bytes) -> str:
    if not data:
        raise ValueError("empty payload")
    if data[:4] == b"PK\x03\x04" and (
        b"xl/workbook.xml" in data or b"[Content_Types]" in data
    ):
        return XLSX_MIME
    # Reject the obvious non-table magic bytes up front (PDF, PNG, JPEG, etc.)
    rejected_magic = (
        b"%PDF-",
        b"\x89PNG",
        b"\xff\xd8\xff",
        b"GIF87a",
        b"GIF89a",
        b"\x00\x00\x00",
    )
    for m in rejected_magic:
        if data.startswith(m):
            raise ValueError("unsupported table mime")
    head = data[:1024].decode("utf-8", errors="ignore")
    if not head:
        raise ValueError("unsupported table mime")
    has_separator = any(d in head for d in (",", ";", "\t"))
    has_newline = "\n" in head or "\r" in head
    # Multi-column CSV is the easy case.
    if has_separator and (has_newline or len(head) < 1024):
        return CSV_MIME
    # Single-column CSV: rows separated by newlines, no commas. Accept
    # ASCII/UTF-8 printable text + a newline, but reject obvious binary
    # (a NUL byte anywhere in the first KB).
    if has_newline and b"\x00" not in data[:1024]:
        return CSV_MIME
    raise ValueError("unsupported table mime")


def read_table(data: bytes, mime: str, sheet_name: str | None = None) -> pd.DataFrame:
    """Read CSV/XLSX bytes into a DataFrame.

    For XLSX, ``sheet_name`` selects which worksheet to load. When ``None``,
    the workbook's active (first) sheet is used — preserves pre-multi-sheet
    behaviour for legacy datasets that don't carry sheet metadata.
    """
    if mime == CSV_MIME:
        return pd.read_csv(io.BytesIO(data))
    if mime == XLSX_MIME:
        wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        if sheet_name is not None and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return pd.DataFrame()
        header = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
        body = rows[1:]
        df = pd.DataFrame(body, columns=header)
        return _coerce_obvious_numerics(df)
    raise ValueError(f"unsupported mime {mime}")


def read_dataset(data: bytes, dataset) -> pd.DataFrame:
    """Convenience: read a dataset's bytes honouring its ``sheet_name``
    metadata (for multi-sheet XLSX). Falls back to the active sheet."""
    meta = getattr(dataset, "dataset_metadata", None) or {}
    sheet = None
    if isinstance(meta, dict):
        sheet = meta.get("sheet_name")
        if not isinstance(sheet, str):
            sheet = None
    return read_table(data, dataset.file_type, sheet_name=sheet)


def list_xlsx_sheets(data: bytes) -> list[str]:
    """Return the sheet names of an XLSX workbook in workbook order.

    Returns an empty list if the input is not a valid XLSX. Never raises.
    """
    try:
        wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    except Exception:
        return []
    return list(wb.sheetnames)


_SUBJECT_NAME_RE = re.compile(r"(patient|subject|id|pid|case)", re.I)
_TIME_LIKE_NAME_RE = re.compile(
    r"(time|visit|wave|week|month|year|tp|timepoint|followup|fu)",
    re.I,
)


def detect_long_format(df: pd.DataFrame) -> dict | None:
    """Heuristic: does the dataframe look like long-format repeated measures?

    Requires (a) a subject-id-shaped column (name hint OR cardinality at
    least half the rows but with average 2+ repeats), (b) a time/wave-like
    column with low cardinality (2-20 levels), and (c) at least one numeric
    outcome column besides the two. Returns ``None`` for wide-format or
    cross-sectional data; never raises.
    """
    try:
        if df is None or len(df) < 6 or df.shape[1] < 3:
            return None
        n_rows = int(len(df))
        # Two-pass: prefer name-matched candidates (patient_id, subject_id …).
        # Only fall back to cardinality-based scan if none found.
        candidates: list[tuple[str, int, bool]] = []
        for col in df.columns:
            try:
                nunique = int(df[col].nunique(dropna=True))
            except Exception:
                continue
            if nunique < 3 or nunique >= n_rows:
                continue
            avg_per_id = n_rows / nunique
            if avg_per_id < 2 or avg_per_id > 20:
                continue
            name_match = bool(_SUBJECT_NAME_RE.search(str(col)))
            tight = abs(avg_per_id - round(avg_per_id)) < 0.05
            if name_match or tight:
                candidates.append((str(col), nunique, name_match))
        if not candidates:
            return None
        # Pick the best: name-matched > more subjects.
        candidates.sort(key=lambda x: (not x[2], -x[1]))
        subject_col, best_n_subjects, _ = candidates[0]
        # Candidate time-like column.
        time_col: str | None = None
        for col in df.columns:
            if col == subject_col:
                continue
            try:
                nunique = int(df[col].nunique(dropna=True))
            except Exception:
                continue
            if 2 <= nunique <= 20 and _TIME_LIKE_NAME_RE.search(str(col)):
                time_col = str(col)
                break
        # Must have a numeric outcome besides the two.
        has_numeric = any(
            pd.api.types.is_numeric_dtype(df[c])
            for c in df.columns
            if c != subject_col and c != time_col
        )
        if not has_numeric:
            return None
        return {
            "subject_col": subject_col,
            "time_col": time_col,
            "n_per_subject": round(n_rows / best_n_subjects, 2),
            "n_subjects": best_n_subjects,
        }
    except Exception:
        return None


def _coerce_obvious_numerics(df: pd.DataFrame) -> pd.DataFrame:
    for col in df.columns:
        s = df[col]
        if s.dtype == object:
            non_null = s.dropna()
            if len(non_null) > 0 and all(isinstance(v, (int, float, np.integer, np.floating)) for v in non_null):
                df[col] = pd.to_numeric(s, errors="ignore")
    return df


def _looks_ordinal(values: list[str]) -> bool:
    lowered = [v.strip().lower() for v in values]
    if all(v.isdigit() for v in lowered) and len(set(lowered)) <= 10:
        return True
    if all(v in _ORDINAL_TOKENS for v in lowered):
        return True
    return False


def _stringify(v: object) -> str:
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def _distinct_non_null(series: pd.Series) -> list[object]:
    seen: list[object] = []
    s = set()
    for v in series.tolist():
        if v is None:
            continue
        if isinstance(v, float) and np.isnan(v):
            continue
        key = _stringify(v)
        if key in s:
            continue
        s.add(key)
        seen.append(v)
    return seen


def infer_columns(df: pd.DataFrame) -> list[InferredColumn]:
    out: list[InferredColumn] = []
    for pos, col in enumerate(df.columns):
        series = df[col]
        n_missing = int(series.isna().sum())
        distinct = _distinct_non_null(series)
        sample = [_stringify(v) for v in distinct[:5]]
        inferred = _infer_one(col, series, distinct)
        out.append(
            InferredColumn(
                name=str(col),
                position=pos,
                inferred_type=inferred,
                n_missing=n_missing,
                sample_values=sample,
            )
        )
    return out


def _infer_one(name: object, series: pd.Series, distinct: list[object]) -> str:
    if len(distinct) == 0:
        return "unknown"
    if len(distinct) == 1:
        return "unknown"

    if pd.api.types.is_datetime64_any_dtype(series):
        return "time"

    name_str = str(name)
    if _TIME_NAME_RE.match(name_str):
        sample = series.dropna().astype(str).head(5).tolist()
        if sample and all(_parseable_datetime(s) for s in sample):
            return "time"

    if pd.api.types.is_numeric_dtype(series):
        return "numeric"

    distinct_str = [_stringify(v) for v in distinct]

    if _looks_ordinal(distinct_str):
        return "ordinal"

    avg_len = (
        sum(len(s) for s in distinct_str) / max(1, len(distinct_str)) if distinct_str else 0
    )
    if len(distinct) <= 10 and avg_len <= 25:
        return "nominal"

    return "nominal"


def _parseable_datetime(s: str) -> bool:
    try:
        pd.to_datetime(s, errors="raise")
        return True
    except (ValueError, TypeError):
        return False


def ingest(data: bytes, mime: str, sheet_name: str | None = None) -> IngestResult:
    df = read_table(data, mime, sheet_name=sheet_name)
    columns = infer_columns(df)
    return IngestResult(
        n_rows=int(df.shape[0]),
        n_columns=int(df.shape[1]),
        columns=columns,
        long_format_hint=detect_long_format(df),
    )
